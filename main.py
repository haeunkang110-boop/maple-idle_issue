"""
메이플키우기 이슈 히스토리 수집기
- 모드 1 (자동): 나무위키 이슈 파싱 + DC 동향 분석
- 모드 2 (수동): 직접 입력한 이슈 내용 + DC 동향 분석 → 엑셀에 행 추가
"""

import os, re, json, time, datetime, sys
import requests
from bs4 import BeautifulSoup
from google import genai
from google.genai import types
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 설정 ──────────────────────────────────────
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
ISSUE_INPUT    = os.environ.get("ISSUE_INPUT", "")   # 수동 모드: Run workflow 입력값
OUTPUT_FILE    = "meki_이슈_히스토리.xlsx"

NAMU_URL  = "https://namu.wiki/w/%EB%A9%94%EC%9D%B4%ED%94%8C%20%ED%82%A4%EC%9A%B0%EA%B8%B0"
DC_URL    = "https://gall.dcinside.com/mgallery/board/lists?id=maplerpg"
DC_SEARCH = "https://gall.dcinside.com/mgallery/board/lists?id=maplerpg&s_type=search_subject_memo&s_keyword={kw}"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9",
    "Referer": "https://gall.dcinside.com",
}

# ── 색상 ──────────────────────────────────────
COLOR_TITLE  = "1B2A4A"
COLOR_HEADER = "C8A870"
COLOR_FONT   = "FFFFFF"
COLOR_ODD    = "F9F9F9"
COLOR_EVEN   = "FFFFFF"

def _fill(c): return PatternFill("solid", fgColor=c)
def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _hfont(): return Font(name="Arial", bold=True, color=COLOR_FONT, size=10)
def _bfont(): return Font(name="Arial", size=10)


# ── DC인사이드 수집 ────────────────────────────
def fetch_dc_posts(keyword: str = "", max_posts: int = 20) -> list[dict]:
    """DC 갤러리 최신 글 또는 키워드 검색 결과 수집"""
    url = DC_SEARCH.format(kw=requests.utils.quote(keyword)) if keyword else DC_URL
    try:
        resp = requests.get(url, headers=HEADERS, timeout=10)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        posts = []
        for row in soup.select("tr.ub-content")[:max_posts]:
            title_el = row.select_one("td.gall_tit a:not(.reply_num)")
            date_el  = row.select_one("td.gall_date")
            rec_el   = row.select_one("td.gall_recommend")
            if not title_el:
                continue
            posts.append({
                "title":  title_el.get_text(strip=True),
                "date":   date_el.get_text(strip=True) if date_el else "",
                "recomm": rec_el.get_text(strip=True) if rec_el else "0",
            })
        return posts
    except Exception as e:
        print(f"[WARN] DC 수집 실패: {e}")
        return []


# ── 나무위키 이슈 파싱 ────────────────────────
def fetch_namuwiki_issues() -> list[dict]:
    """나무위키 메키 문서에서 이슈/논란 섹션 파싱"""
    try:
        resp = requests.get(NAMU_URL, headers={**HEADERS, "Referer": "https://namu.wiki"}, timeout=15)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        issues = []
        # 이슈/논란 관련 섹션 찾기
        for heading in soup.find_all(["h2", "h3", "h4"]):
            text = heading.get_text(strip=True)
            if any(kw in text for kw in ["논란", "이슈", "오류", "사건", "버그", "보상", "패치"]):
                content_parts = []
                for sibling in heading.find_next_siblings():
                    if sibling.name in ["h2", "h3", "h4"]:
                        break
                    t = sibling.get_text(separator=" ", strip=True)
                    if t:
                        content_parts.append(t)
                content = " ".join(content_parts)[:1500]
                if content:
                    issues.append({"title": text, "content": content})
        return issues
    except Exception as e:
        print(f"[WARN] 나무위키 수집 실패: {e}")
        return []


# ── Gemini 분석 ───────────────────────────────
def analyze_with_gemini(issue_text: str, dc_posts: list[dict]) -> dict:
    client = genai.Client(api_key=GEMINI_API_KEY)

    dc_summary = "\n".join([f"- [{p['recomm']}추천] {p['title']}" for p in dc_posts[:15]]) or "없음"

    prompt = f"""
아래는 메이플키우기(메키) 이슈 내용과 DC인사이드 갤러리 반응입니다.
JSON 형식으로만 답하세요. 백틱이나 다른 텍스트 없이 순수 JSON만 반환하세요.

[이슈 내용]
{issue_text[:2000]}

[DC인사이드 반응 (추천수 포함)]
{dc_summary}

반환 JSON:
{{
  "이슈_구분": "수치 오류 | 서버 오류 | UI 버그 | 밸런스 문제 | 결제 오류 | 확률 오류 | 운영 논란 | 기타 중 하나",
  "이슈_핵심_요약": "30자 이내",
  "이슈_상세": "2~3문장",
  "대응_내용": "패치/롤백/공지 등 대응 내용",
  "보상_내용": "보상 항목과 수량 (없으면 없음)",
  "주요_동향": "납득 | 긍정 | 부정 | 혼재 중 하나",
  "비고": "동향 판단 근거 한 문장"
}}
"""
    try:
        response = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=prompt,
        )
        text = response.text.strip().lstrip("```json").lstrip("```").rstrip("```").strip()
        return json.loads(text)
    except Exception as e:
        print(f"[WARN] Gemini 분석 실패: {e}")
        return {
            "이슈_구분": "기타", "이슈_핵심_요약": issue_text[:30],
            "이슈_상세": issue_text[:200], "대응_내용": "",
            "보상_내용": "", "주요_동향": "", "비고": "",
        }


# ── 엑셀 로드 또는 생성 ───────────────────────
def load_or_create_wb(path: str):
    if os.path.exists(path):
        return load_workbook(path), False
    wb = Workbook()
    ws = wb.active
    ws.title = "이슈 히스토리"

    # 타이틀
    ws.merge_cells("A1:I1")
    ws["A1"] = "메이플 키우기 이슈 히스토리"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color=COLOR_FONT)
    ws["A1"].fill = _fill(COLOR_TITLE)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        ("No.", 5), ("발생일", 12), ("이슈 구분", 18),
        ("이슈 핵심 요약", 36), ("이슈 상세", 44),
        ("대응 내용", 36), ("보상 내용", 28),
        ("주요 동향", 16), ("비고", 28),
    ]
    for col_idx, (title, width) in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx, value=title)
        c.font = _hfont()
        c.fill = _fill(COLOR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"
    return wb, True


def append_rows(wb, rows: list[dict]):
    ws = wb.active
    # 현재 마지막 행 번호 파악
    last_row = ws.max_row
    if last_row < 2:
        last_row = 2
    start_no = last_row - 1  # 헤더 2행 제외한 데이터 수

    for i, row in enumerate(rows):
        r = last_row + 1 + i
        no = start_no + i + 1
        bg = COLOR_ODD if no % 2 == 1 else COLOR_EVEN
        values = [
            no,
            row.get("date", datetime.date.today().strftime("%Y-%m-%d")),
            row.get("이슈_구분", ""),
            row.get("이슈_핵심_요약", ""),
            row.get("이슈_상세", ""),
            row.get("대응_내용", ""),
            row.get("보상_내용", ""),
            row.get("주요_동향", ""),
            row.get("비고", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.fill = _fill(bg)
            c.border = _border()
            c.font = _bfont()
            c.alignment = Alignment(
                horizontal="center" if col_idx in (1, 2, 3, 8) else "left",
                vertical="center", wrap_text=True,
            )
        ws.row_dimensions[r].height = 52


# ── 메인 ──────────────────────────────────────
def main():
    if not GEMINI_API_KEY:
        raise ValueError("GEMINI_API_KEY 환경변수가 없습니다.")

    wb, is_new = load_or_create_wb(OUTPUT_FILE)
    new_rows = []

    if ISSUE_INPUT:
        # ── 모드 2: 수동 입력 ──────────────────
        print(f"=== 수동 모드: '{ISSUE_INPUT}' ===")

        # 입력 키워드로 DC 검색
        keywords = ISSUE_INPUT.split()[:3]
        dc_posts = []
        for kw in keywords:
            dc_posts += fetch_dc_posts(keyword=kw, max_posts=10)
            time.sleep(0.5)
        # 중복 제거
        seen = set()
        dc_unique = []
        for p in dc_posts:
            if p["title"] not in seen:
                seen.add(p["title"])
                dc_unique.append(p)

        print(f"  DC 수집: {len(dc_unique)}건")
        analysis = analyze_with_gemini(ISSUE_INPUT, dc_unique)
        new_rows.append(analysis)

    else:
        # ── 모드 1: 자동 수집 (나무위키) ───────
        print("=== 자동 모드: 나무위키 수집 ===")
        issues = fetch_namuwiki_issues()
        print(f"  나무위키 이슈 섹션: {len(issues)}건")

        # DC 최신 글 수집 (전체 동향용)
        dc_posts = fetch_dc_posts(max_posts=30)
        print(f"  DC 최신 글: {len(dc_posts)}건")

        for idx, issue in enumerate(issues):
            print(f"  [{idx+1}/{len(issues)}] {issue['title'][:30]}")
            # 이슈 키워드로 DC에서 관련 글 필터링
            kws = issue["title"].replace(" ", "")[:5]
            related_dc = [p for p in dc_posts if kws[:3] in p["title"]] or dc_posts[:10]

            analysis = analyze_with_gemini(
                f"제목: {issue['title']}\n\n{issue['content']}", related_dc
            )
            new_rows.append(analysis)
            time.sleep(1)

    append_rows(wb, new_rows)
    wb.save(OUTPUT_FILE)
    print(f"\n[OK] {OUTPUT_FILE} 저장 완료 ({len(new_rows)}건 추가)")


if __name__ == "__main__":
    main()
